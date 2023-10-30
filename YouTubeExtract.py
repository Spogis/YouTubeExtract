# -*- coding: utf-8 -*-
"""
Created on Fri Apr  9 15:06:32 2021

@author: Nicolas Spogis
"""

# importar os pacotes necessários
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import matplotlib.pyplot as plt
from wordcloud import WordCloud, STOPWORDS
from selenium import webdriver
import time
import datetime

global MinChar
global chrome_path
MinChar = 2
global MaxHeight
MaxHeight = 20

chrome_path = r'C:/bin/chromedriver.exe'

def GetURLsFromYoutube(Busca):
    # Open The Output Excel
    global chrome_path
    if os.path.exists("./SiteLists/URL_List.xlsx"):
        os.remove("./SiteLists/URL_List.xlsx")
    
    wb = Workbook()
    wb.save(filename = './SiteLists/URL_List.xlsx')
    workbook = load_workbook(filename="./SiteLists/URL_List.xlsx")
    sheet = workbook.active
    
    for url in Busca:
        # Start the driver
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--mute-audio")
        driver = webdriver.Chrome(executable_path=chrome_path, options=chrome_options)
    
        driver.get(url)
        time.sleep(5)
        height = driver.execute_script("return document.documentElement.scrollHeight")
        lastheight = 0
        
        countheight=0
        while True:
            countheight = countheight + 1
            if countheight == MaxHeight:
                break
            if lastheight == height:
                break
            lastheight = height
            driver.execute_script("window.scrollTo(0, " + str(height) + ");")
            time.sleep(2)
            height = driver.execute_script("return document.documentElement.scrollHeight")
        
        user_data = driver.find_elements_by_xpath('//*[@id="video-title"]')
        
        for i in user_data:
            TempList = (i.get_attribute('href'))
            rows = sheet.max_row
            sheet.cell(row=rows+1, column=1).value = TempList
        
        driver.quit()
    
    workbook.save("./SiteLists/URL_List.xlsx")
    workbook.close()
          
def getSiteList(open_file):
    global SitesURLs
    open_file="./SiteLists/"+open_file
    workbook = load_workbook(open_file)
    sheet = workbook.active
    SitesURLs =[]

    for cell in sheet['A']:
        if cell.value is not None:
            SitesURLs.append(cell.value)
            
    workbook.close()

def getTags():
    global tags
    # Open The Output Excel
    if os.path.exists("HData.xlsx"):
        os.remove("HData.xlsx")

    wb = Workbook()
    wb.save(filename = 'HData.xlsx')
    workbook = load_workbook(filename="HData.xlsx")
    sheet = workbook.active

    for url in SitesURLs:
        request = requests.get(url)
        html = BeautifulSoup(request.content, "html.parser")
        tags = html.find_all("meta",property="og:video:tag")
    
        for tag in tags:
            rows = sheet.max_row
            if len(tag['content'])>MinChar: 
                sheet.cell(row=rows+1, column=1).value = tag['content']
            
    workbook.save("HData.xlsx")
    workbook.close()

def getOtherData():
    global tags
    global title
    global UrlData
    global description
    
    # Open The Output Excel
    if os.path.exists("VideoData.xlsx"):
        os.remove("VideoData.xlsx")

    wb = Workbook()
    wb.save(filename = 'VideoData.xlsx')
    workbook = load_workbook(filename="VideoData.xlsx")
    sheet = workbook.active

    for url in SitesURLs:
        request = requests.get(url)
        html = BeautifulSoup(request.content, "html.parser")
        title = html.find("meta", property="og:title")
        UrlData = html.find("meta", property="og:url")
        description = html.find("meta",property="og:description")
            
        rows = sheet.max_row
        sheet.cell(row=rows+1, column=1).value = title.get('content')
        sheet.cell(row=rows+1, column=2).value = UrlData.get('content')
        sheet.cell(row=rows+1, column=3).value = description.get('content')
            
    workbook.save("VideoData.xlsx")
    workbook.close()

def GenerateWordCloud(NomeDaFigura):
    
    # lista de stopword
    STOPWORDS_DATA = []
    workbook = load_workbook(filename="./Others/STOPWORDS.xlsx")
    sheet = workbook.active
    for cell in sheet['A']:
        STOPWORDS_DATA.append(cell.value)
        
    workbook.close()
    
    stopwords = set(STOPWORDS)
    stopwords.update(STOPWORDS_DATA)

    # Start by opening the spreadsheet and selecting the main sheet
    workbook = load_workbook(filename="HData.xlsx")
    sheet = workbook.active
    summary =[]
    
    for cell in sheet['A']:
        summary.append(cell.value)
    
    # concatenar as palavras
    all_summary = ' '.join([str(elem) for elem in summary])
    
    # gerar uma wordcloud
    wc = WordCloud(stopwords=stopwords,prefer_horizontal=.1,contour_width=5, font_step=2,
                          width=800, height=600, max_words = 200, max_font_size=40)
    # generate word cloud
    wc.generate(all_summary)
     
    # mostrar a imagem final
    fig, ax = plt.subplots(figsize=(1200,1200))
    ax.imshow(wc, interpolation='bilinear')
    ax.set_axis_off()
     
    plt.imshow(wc, interpolation='bilinear');
    
    if NomeDaFigura == "Incremented":
        # Determine incremented filename
        a= getNextFilePath("./Pictures")
        filename = "./Pictures/WordCloud_" +str(a) + ".png"
        wc.to_file(filename)
    else:
        dt = datetime.datetime.today()
        filename = "./Pictures/" + str(dt.day) + "_" + str(dt.month) + "_" + str(dt.year) + "-" +str(NomeDaFigura) + ".png"
        wc.to_file(filename)

def getNextFilePath(output_folder):
    highest_num = 0
    for f in os.listdir(output_folder):
        highest_num = highest_num + 1

    output_file = str(highest_num+1)
    return output_file

def RunAll():
    GetURLsFromYoutube(BuscaSites)
    getSiteList("URL_List.xlsx")
    getTags()
    GenerateWordCloud(NomeDaFigura)


#BuscaSites = ["https://www.youtube.com/results?search_query=ASPEN+Plus&sp=CAMSBAgFEAE%253D",
#              "https://www.youtube.com/results?search_query=DWSIM&sp=CAMSBAgFEAE%253Dhttps://www.youtube.com/results?search_query=DWSIM&sp=CAMSBAgFEAE%253D",
#              "https://www.youtube.com/results?search_query=Hysys&sp=CAMSBAgFEAE%253D"] 
#RunAll()

#Aprender
#BuscaSites = ["https://www.youtube.com/channel/UCtFRv9O2AHqOZjjynzrv-xg"]
#RunAll()

#Marcelo Barros
#BuscaSites = ["https://www.youtube.com/channel/UCzkg1HCRcmWR-7ZX2sS5LtQ/videos"]
#RunAll()

#FBGA oficial
#BuscaSites = ["https://www.youtube.com/c/FBGAoficial/videos"]
#RunAll()

#NOCAUTE - Blog do Fernando Morais
#BuscaSites = ["https://www.youtube.com/c/NocauteTV/videos"]
#RunAll()

#IBREI
#BuscaSites = ["https://www.youtube.com/channel/UCyP83xHyFqWA7rhxcCQuLQA/videos"]
#RunAll()

#Finocchio & Ustra
#BuscaSites = ["https://www.youtube.com/channel/UCBzARAQulJ4JsPm7iyPVfXg/videos"]
#RunAll()

#Casca Grossa Suprema
#BuscaSites = ["https://www.youtube.com/c/CascaGrossaSuprema/videos"]
#RunAll()

# NomeDaFigura = "Notícias sobre negócios"
# BuscaSites = ["https://www.youtube.com/channel/UCQxtLzG3ckdfsnxd8MDfx2g"]
# RunAll()

# NomeDaFigura = "Notícias sobre ciência e tecnologia"
# BuscaSites = ["https://www.youtube.com/channel/UCZ7QHqCYPE3Zi2Tt1sBobEw"]
# RunAll()

# NomeDaFigura = "Notícias mundiais"
# BuscaSites = ["https://www.youtube.com/channel/UCvAvFl2OGsuDSoOo93Kd0nA"]
# RunAll()

# NomeDaFigura = "Notícias nacionais"
# BuscaSites = ["https://www.youtube.com/channel/UCcE169gw8kJCzyCJZXb7DQw"]
# RunAll()

# NomeDaFigura = "Notícias sobre saúde"
# BuscaSites = ["https://www.youtube.com/channel/UCn371zWk5jljg-ycIXkEUSA"]
# RunAll()

# NomeDaFigura = "Notícias"
# BuscaSites = ["https://www.youtube.com/channel/UCYfdidRxbB8Qhf0Nx7ioOYw"]
# RunAll()

# NomeDaFigura = "Ao Vivo"
# BuscaSites = ["https://www.youtube.com/channel/UC4R8DWoMoI7CAwX8_LjQHig"]
# RunAll()

# NomeDaFigura = "Em Alta"
# BuscaSites = ["https://www.youtube.com/feed/trending"]
# RunAll()

# NomeDaFigura = "Popular on YouTube - Worldwide"
# BuscaSites = ["https://www.youtube.com/channel/UCgGzSIa8zIsJHbSs0bLplag"]
# RunAll()

# NomeDaFigura = "Global News"
# BuscaSites = ["https://www.youtube.com/channel/UChLtXXpo4Ge1ReTEboVvTDg"]
# RunAll()

# NomeDaFigura = "BETAEQ"
# BuscaSites = ["https://www.youtube.com/c/BetaEQ/videos"]
# RunAll()

#NomeDaFigura = "Noticias da Semana"
#BuscaSites = ["https://www.youtube.com/channel/UCQxtLzG3ckdfsnxd8MDfx2g",
#              "https://www.youtube.com/channel/UCZ7QHqCYPE3Zi2Tt1sBobEw",
#              "https://www.youtube.com/channel/UCvAvFl2OGsuDSoOo93Kd0nA",
#              "https://www.youtube.com/channel/UCcE169gw8kJCzyCJZXb7DQw",
#              "https://www.youtube.com/channel/UCn371zWk5jljg-ycIXkEUSA",
#              "https://www.youtube.com/channel/UCYfdidRxbB8Qhf0Nx7ioOYw",
#              "https://www.youtube.com/channel/UC4R8DWoMoI7CAwX8_LjQHig",
#              "https://www.youtube.com/feed/trending",
#              "https://www.youtube.com/channel/UCgGzSIa8zIsJHbSs0bLplag",
#              "https://www.youtube.com/channel/UChLtXXpo4Ge1ReTEboVvTDg"]
#RunAll()

NomeDaFigura = "Canal CascaGrossaSuprema"
BuscaSites = ["https://www.youtube.com/c/CascaGrossaSuprema/videos"]
RunAll()